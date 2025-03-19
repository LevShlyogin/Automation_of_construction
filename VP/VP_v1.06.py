#___________________________________________________________
#ПОДКЛЮЧАЕМ НЕОБХОДИМЫЕ БИБЛИОТЕКИ
#___________________________________________________________

import tkinter as tk
from tkinter import filedialog
import os
import openpyxl
import xlsxwriter
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Side, Border
from openpyxl.styles import Alignment
from seuif97 import *

#___________________________________________________________
# ИСХОДНЫЕ ДАННЫЕ
#___________________________________________________________

#Тип баланса a
type_balance_regeneration_a = 2
name_balance_type_a = 'a' # Старый баланс
name_extraction_a = 'Отборы, и др.:G-расход т/час;H-энтальпия ккал/кг;Р кг/см2;Т град'
name_regeneration_a = 'Участки регенерации и подогрева сетевой воды :расход, т/час; энтальпия, ккал/кг'
name_water_heating_a = 'Подогрев сетевой воды: температура, град; недогрев, град; давление пара, ата'

compartment_name_value_list_a = ['Gотс','P1','P2','h1','hад','hи','t1','x2','GV2','кпд','кпд*','Nотс','t2']
compartment_name_unit_list_a = ['т/ч','кгс/см2','кгс/см2','ккал/кг','ккал/кг','ккал/кг','°С','-','м3/с','%','%','МВт','°С']

main_parameters_list_a = ['B8','F8','H8','I8','J8','K8','L8','M8','N8','G8']
main_parameters_value_list_a = ['Gтурб','Ne','Qтурб','Qот','Wсв','tоб','tпр','tохл','hпв','qe']
main_parameters_unit_list_a = ['т/ч','МВт','Гкал/ч','Гкал/ч','м3/ч','°С','°С','°С','ккал/кг','ккал/кВтч']

#Тип баланса b
type_balance_regeneration_b = 3
name_balance_type_b = 'b' # Новый баланс
name_extraction_b = 'Отборы,конденсат,уплотнения и др.:G-т/час;H-ккал/кг;Р кг/см2;Т град'
name_regeneration_b = 'Участки регенерации и подогрева сетевой воды:G т/ч; H ккал/ч; P кг/см2'
name_water_heating_b = 'Подогрева сетевой воды: Т, град; недогрев, град; давление пара, ата'

compartment_name_value_list_b = ['Gотс', 'P1', 'P2', 'h1', 'h2', 'hад', 't1', 't2', 'x2', 'GV2', 'кпд', 'кпд*', 'Nотс']
compartment_name_unit_list_b = ['т/ч', 'кгс/см2', 'кгс/см2', 'ккал/кг', 'ккал/кг', 'ккал/кг', '°С', '°С', '-', 'м3/с', '%', '%', 'МВт']

main_parameters_list_b = ['B8', 'F8', 'H8', 'I8', 'J8', 'K8', 'L8', 'M8', 'N8', 'O8','G8']
main_parameters_value_list_b = ['Gтурб', 'Ne', 'Qтурб', 'Qот', 'Wсв', 'tоб', 'tпр', 'tохл', 'tпв', 'hпв','qe']
main_parameters_unit_list_b = ['т/ч', 'МВт', 'Гкал/ч', 'Гкал/ч', 'м3/ч', '°С', '°С', '°С', '°С', 'ккал/кг','ккал/кВтч']

#Общие
main_parameters_list_all = ['TBp0','TBt0','Tbtpp','Tbg00','Tbpp1','Tbtpg1','Tbpg1','TBtox','TBt1w','TBwset','TBkorNe','Tbch1p','TBch1','TBch2','TBch3']
main_parameters_value_list_all = ['P0','T0','Tпп','G0','Pнд','Tнд','Gнд','tов','tоб','Wсв','dNe','Pсн1','Gсн1','Gсн2','Gсн3']
main_parameters_unit_list_all = ['кгс/см2','°C','°C','т/ч','кгс/см2','°C','т/ч','°C','°C','т/ч','-','кгс/см2','т/ч','т/ч','т/ч']

extraction_name_value_list = ['G', 'H', 'P', 'T', 'v', 'Gv2']
extraction_name_unit_list = ['т/ч', 'ккал/кг', 'кгс/см2', '°С', 'м3/кг', 'м3/с']
regeneration_name_value_list = ['G', 'H', 'P', 'T']
regeneration_name_unit_list = ['т/ч', 'ккал/кг', 'кгс/см2', '°С']
water_heating_name_value_list = ['T', 'dT', 'Pот']
water_heating_name_unit_list = ['°С', '°С', 'кгс/см2']

speed_name_value_list = ['','Z', 'Dу', '', 'F', 'Сотб', '','R1','R2','Cщели','Fщели', 'Lщели', '','Z=1','Z=2','Z=3','Z=4']
speed_name_unit_list = ['','шт', 'мм', '', 'см2', 'м/с', '','мм','мм','м/с','см2','мм','','шт','шт','шт','шт']

name_compartment = 'Отсек'

#_______________________________________________________________________________________
# ФУНКЦИИ
#_______________________________________________________________________________________

# Функция для поиска ячеек с заданным словом и изменением координат ячейки
def find_words_in_excel(file_path, word, num_row_start, num_col_start):
    
    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Назначаем базовые координаты ячейки
    base_row = None
    base_col = None
    
    # Ищем ячейки с заданным словом
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == word:
                # Определяем строку и колонку найденной ячейки
                base_row = cell.row
                base_col = cell.column
                break
        
        if base_row is not None and base_col is not None:
            break
    # Изменяем строку и колонку найденной ячейки при введении отрицательного значения функкция работает
    if base_row is not None and base_col is not None:
        new_row = base_row + num_row_start
        new_col = base_col + num_col_start
        new_cell = sheet.cell(row=new_row, column=new_col)
    
    # Возврощаем нужную координату ячейки
    return new_cell

#_______________________________________________________________________________________
#_______________________________________________________________________________________

# Функция для поиска ячеек с заданным словом и изменением координат ячейки
def find_words_in_excel_2(file_path, word, num_row_start, num_col_start):
    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Назначаем базовые координаты ячейки
    base_row = None
    base_col = None
    
    # Ищем ячейки с заданным словом
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == word:
                # Определяем строку и колонку найденной ячейки
                base_row = cell.row
                base_col = cell.column
                break
        
        if base_row is not None and base_col is not None:
            break
    # Изменяем строку и колонку найденной ячейки при введении отрицательного значения функкция работает
    if base_row is not None and base_col is not None:
        new_row = num_row_start
        new_col = base_col + num_col_start
        new_cell = sheet.cell(row=new_row, column=new_col)
    
    # Возвращаем нужную координату ячейки
    return new_cell

#_______________________________________________________________________________________
#_______________________________________________________________________________________

# Функция для изменением координат ячейки
def new_cell_to_excel(file_path, coordinate, num_row, num_col): # Координата формата "А1", "N45" и т.п. 
    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Определяем ячейку на листе по заданной координате
    cell_coordinate = sheet[coordinate]
    
    # Определяем координаты строку и колонку заданной ячейки
    cell_coordinate_row = cell_coordinate.row
    cell_coordinate_col = cell_coordinate.column
    
    # Меняем координаты ячейки
    new_row = cell_coordinate_row + num_row
    new_col = cell_coordinate_col + num_col
    new_cell = sheet.cell(row=new_row, column=new_col)
    
    return new_cell.coordinate

#_______________________________________________________________________________________
#_______________________________________________________________________________________

# Функция для обработки области отсеков
def process_excel_compartment(file_path, start_row, end_row, start_col):
    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # Создаем словарь для хранения результатов
    result_dict = {}
    #Заполнение словаря параметрами отсека
    for i in range(start_row, end_row+1):
        compartment_list = []
        for j in range(start_col+1, 16):
            compartment_list.append(sheet.cell(row=i, column=j).value)
            result_dict[sheet.cell(row=i, column=start_col).value] = compartment_list
    #Закрываем файл после чтения
    workbook.close()
    
    # Возвращаем результат
    return result_dict
#_______________________________________________________________________________________
#_______________________________________________________________________________________

# Функция для обработки основных параметров
def process_excel_main(file_path, parameters_list):
    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # Создаем словарь для хранения результатов
    result_dict = []
    #Заполнение словаря параметрами отсека
    for i in parameters_list:
        cell_value = sheet[i].value
        result_dict.append(cell_value)
    #Закрываем файл после чтения
    workbook.close()
    
    # Возвращаем результат
    return result_dict

#_______________________________________________________________________________________
#_______________________________________________________________________________________

# Функция для определения типа баланса
def type_balance(file_path):
    # Открываем Excel файл
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    # Тип баланса 1
    balance_type = name_balance_type_a
    type_balance_regeneration = type_balance_regeneration_a
    name_extraction = name_extraction_a
    name_regeneration = name_regeneration_a
    name_water_heating = name_water_heating_a
    compartment_name_value_list = compartment_name_value_list_a
    compartment_name_unit_list = compartment_name_unit_list_a
    main_parameters_list = main_parameters_list_a 
    main_parameters_value_list = main_parameters_value_list_a
    main_parameters_unit_list = main_parameters_unit_list_a

    # Проходим по всем строкам и ищем совпадение
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == name_extraction_b:
                balance_type = name_balance_type_b
                type_balance_regeneration = type_balance_regeneration_b
                name_extraction = name_extraction_b
                name_regeneration = name_regeneration_b
                name_water_heating = name_water_heating_b
                compartment_name_value_list = compartment_name_value_list_b
                compartment_name_unit_list = compartment_name_unit_list_b
                main_parameters_list = main_parameters_list_b 
                main_parameters_value_list = main_parameters_value_list_b
                main_parameters_unit_list = main_parameters_unit_list_b
                break
    return balance_type, name_extraction, name_regeneration, name_water_heating, compartment_name_value_list, compartment_name_unit_list, main_parameters_list, main_parameters_value_list, main_parameters_unit_list, type_balance_regeneration

#_______________________________________________________________________________________
#_______________________________________________________________________________________

def merge_dictionaries(input_dict):
    # Создаем копию исходного словаря, чтобы не изменять его напрямую
    result_dict = dict(input_dict)

    # Создаем словарь для группировки ключей по значениям 'H' и 'P'
    group_dict = {}

    # Перебираем все ключи для группировки
    for key in input_dict:
        h_value = next((item[1] for item in input_dict[key] if item[0] == 'H'), None)
        p_value = next((item[1] for item in input_dict[key] if item[0] == 'P'), None)
        h_value = f"{h_value:.2f}" if h_value is not None else None
        p_value = f"{p_value:.2f}" if p_value is not None else None

        # Формируем ключ для группировки
        group_key = (h_value, p_value)

        # Добавляем ключ в группу
        if group_key in group_dict:
            group_dict[group_key].append(key)
        else:
            group_dict[group_key] = [key]

    # Перебираем группы для создания новых ключей
    for group_key, keys in group_dict.items():
        if len(keys) > 1:
            new_key = " + ".join(keys)

            # Создаем новый список параметров для нового ключа
            new_values = {}

            # Перебираем параметры из всех ключей в группе
            for key in keys:
                for item in input_dict[key]:
                    param, value = item
                    if param in new_values:
                        if param == 'G':
                            new_values[param] += value
                        else:
                            new_values[param] = value
                    else:
                        new_values[param] = value

            # Преобразуем словарь параметров в список
            new_values_list = [[param, value] for param, value in new_values.items()]

            # Добавляем новый ключ в результирующий словарь
            result_dict[new_key] = new_values_list

    return result_dict

# Функция для обработки области отборов
def process_excel_extraction(file_path, range_str):
    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Заданный диапазон ячеек
    cell_range = sheet[range_str]

    # Создаем словарь для хранения результатов
    result_dict = {}

    # Обрабатываем каждую пару ячеек в заданном диапазоне
    for row in cell_range:
        for i in range(0, len(row), 2):
            if i + 1 < len(row) and row[i].value:
                key_cell_value = str(row[i].value)
                value_cell_value = row[i+1].value                
                
                # Разделяем значение первой ячейки на два фрагмента
                key_fragments = key_cell_value.split(':')
                
                if len(key_fragments) == 2:
                    key = key_fragments[1]
                    if key in result_dict:
                        result_dict[key].append([key_fragments[0], value_cell_value])
                    else:
                        result_dict[key] = [[key_fragments[0], value_cell_value]]

    #Закрываем файл после чтения
    workbook.close()
    for key in result_dict:
        if key != "Ох2" or key != "Ох2 ":
            for key_2, value in parameter_extraction:
                check_and_add(result_dict, key, [key_2, value])
    result_dict = merge_dictionaries(result_dict)

    partial_key = 'Ох2'
    keys_to_delete = [key for key in result_dict if partial_key in key]
    for key in keys_to_delete:
        del result_dict[key]

    #custom_order = ['О', 'С', 'П']
    #def custom_sort_key(key):
    #    for i, letter in enumerate(custom_order):
    #        if key.startswith(letter):
    #            return (0, i)
    #    return (1, key)

    #result_dict = dict(sorted(result_dict.items(), key=lambda item: custom_sort_key(item[0])))
    #result_dict = dict(sorted(result_dict.items()))

    # Расчет температуры, удельного объема и удельного расхода
    for key in result_dict:
        result_dict[key][3][1] = max(0.0, ph(result_dict[key][2][1] * 0.0980665, result_dict[key][1][1] * 4.1868, 1))
        result_dict[key][4][1] = max(0.0, ph(result_dict[key][2][1] * 0.0980665, result_dict[key][1][1] * 4.1868, 3))
        result_dict[key][5][1] = max(0.0, result_dict[key][0][1] * ph(result_dict[key][2][1] * 0.0980665, result_dict[key][1][1] * 4.1868, 3) / 3.6)
    # Возвращаем результат
    return result_dict
#_______________________________________________________________________________________
#_______________________________________________________________________________________

# Функция по добавлению дополнительных ячеек в словарь отбора
def check_and_add(result_dict, key, value):
    found = False
    for i in result_dict.get(key, []):
        if i[0] == value[0]:
            found = True
            break
    if not found:
        result_dict.setdefault(key, []).append(value)

parameter_extraction = [['G', 0.0],['H', 0.0], ['P', 0.0], ['T', 0.0], ['v', 0.0], ['Gv2', 0.0]]

#_______________________________________________________________________________________
#_______________________________________________________________________________________

def process_excel_regeneration(file_path, cell_range, num_cells):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = {}
    
    for row in sheet[cell_range]:
        current_key = None
        values = []
        for cell in row:
            if cell.value is not None:
                if isinstance(cell.value, str):
                    if current_key is not None:
                        data[current_key] = values
                        values = []
                    current_key = cell.value
                else:
                    values.append(cell.value)
        if current_key is not None:
            data[current_key] = values
    
    # Добавляем параметры на линии регенерации
    new_data = {}
    for key, params in data.items():
        if len(params) == 2:
            new_params = params + [0.0, 0.0]
        elif len(params) == 3:
            new_params = params + [0.0]
        elif len(params) == 0:
            new_params = params + [0.0]
        new_data[key] = new_params

    # Расчет температуры ОК/ПВ/СВ
    for key in new_data:
        new_data[key][3] = max(0.0, ph(new_data[key][2] * 0.0980665, new_data[key][1] * 4.1868, 1))

    # Возвращаем результат
    return new_data

def process_excel_water_heating(file_path, cell_range, num_cells):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = {}
    
    for row in sheet[cell_range]:
        current_key = None
        values = []
        for cell in row:
            if cell.value is not None:
                if isinstance(cell.value, str):
                    if current_key is not None:
                        data[current_key] = values
                        values = []
                    current_key = cell.value
                else:
                    values.append(cell.value)
        if current_key is not None:
            data[current_key] = values
    
    # Добавляем параметры на линии регенерации
    new_data = {}
    for key, params in data.items():
        if len(params) == 3:
            new_params = params
        elif len(params) == 0:
            new_params = params + [0.0]
        new_data[key] = new_params
    

    # Возвращаем результат
    return new_data
#_______________________________________________________________________________________
#_______________________________________________________________________________________

def find_empty_rows(filepath, start_coord):
    empty_rows = []
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    max_row_user = ws.max_row + 1
    current_row = start_coord[0]
    current_col = start_coord[1]
    
    count_empty = 0

    while count_empty < 4:
        cell_value = ws.cell(row=current_row, column=current_col).value
        if cell_value is None or cell_value == "":
            empty_rows.append(current_row)
            count_empty += 1
        else:
            count_empty = 0
        
        current_row += 1

        if current_row > max_row_user:
            break
    
    if len(empty_rows) < 4:
        empty_rows += [0] * (4 - len(empty_rows))
    
    return empty_rows

#_______________________________________________________________________________________
#_______________________________________________________________________________________
    
def process_excel_file(path, sheet_number, start_coord):
    # Загрузка excel файла
    wb = load_workbook(path)
    sheet = wb.worksheets[sheet_number - 1]

    # Проверка наличия данных на листе
    if sheet.max_row < start_coord[0] or sheet.max_column < start_coord[1]:
        return # Нет данных на листе, функция завершает выполнение
    
    # Определение количества строк и столбцов данных
    data_rows = sheet.max_row - start_coord[0] + 1
    data_cols = sheet.max_column - start_coord[1] + 1
    
    # Преобразование стартовых координат для использования в циклах
    start_row = start_coord[0]
    start_col = start_coord[1]
    
    # Закрепление области каждого листа на ячейке "B4"
    sheet.freeze_panes = sheet[f"{openpyxl.utils.get_column_letter(2)}{4}"]
    
    #Настройка ширины столбца
    autofit_column(sheet, 1)
    
      
    # Обработка каждого столбца
    for col in range(start_col, start_col + data_cols):
        try:
            column_data = [float(sheet.cell(row=row, column=col).value) for row in range(start_row, start_row + data_rows) if sheet.cell(row=row, column=col).value is not None]
        except (ValueError, TypeError):
            pass
    
        if not column_data:
            continue # Пропустить пустой столбец

        #Настройка ширины столбца
        set_column_width(sheet, col, 7)

        max_value = max([float(value) for value in column_data if value is not None])

        for row in range(start_row, start_row + data_rows):
            cell_value = sheet.cell(row=row, column=col).value

            if cell_value is None:
                continue # Пропустить пустую ячейку

            # Условное форматирование цветом
            fill = PatternFill(start_color='FF9999', end_color='00FF00', fill_type='solid')
            
            # Выделение максимального значения
            if cell_value == max_value:
                # Выделение максимального значения "Жирным" + размер шрифта
                sheet.cell(row=row, column=col).font = Font(bold=True, size=12)
                # Выделение максимального значения "Красным" 
                sheet.cell(row=row, column=col).fill = fill
                max_regim_cell_row = row
            
        # Вставка максимального значения под столбцом
        max_value_cell = sheet.cell(row=start_row + data_rows + 1, column=col)
        max_value_cell.value = max_value
        # Вставка режима на котором достигается максимальное значение.
        max_regim_cell = sheet.cell(row=start_row + data_rows + 2, column=col)
        max_regim_cell.value = sheet.cell(row=max_regim_cell_row, column=1).value
        
        change_text_orientation(sheet, max_regim_cell.coordinate)
        
    # Сохранение изменений
    wb.save(path)
    wb.close()

#_______________________________________________________________________________________
#_______________________________________________________________________________________

# эта функция копия функции выше просто та что выше не воспринимала пустые ячейки, всм тупо пустые столбцы... Эта понимает и работает
def process_excel_file_1(path, sheet_number, start_coord):
    # Загрузка excel файла
    wb = load_workbook(path)
    sheet = wb.worksheets[sheet_number - 1]

    # Проверка наличия данных на листе
    if sheet.max_row < start_coord[0] or sheet.max_column < start_coord[1]:
        return # Нет данных на листе, функция завершает выполнение
    
    # Определение количества строк и столбцов данных
    data_rows = sheet.max_row - start_coord[0] + 1
    data_cols = sheet.max_column - start_coord[1] + 1
    
    # Преобразование стартовых координат для использования в циклах
    start_row = start_coord[0]
    start_col = start_coord[1]
    
    # Закрепление области каждого листа на ячейке "B4"
    sheet.freeze_panes = sheet[f"{openpyxl.utils.get_column_letter(2)}{4}"]
    
    #Настройка ширины столбца
    autofit_column(sheet, 1)
    
    
    # Обработка каждого столбца
    for col in range(start_col, start_col + data_cols):
        column_data = [float(sheet.cell(row=row, column=col).value) for row in range(start_row, start_row + data_rows) if sheet.cell(row=row, column=col).value is not None]
        
        if not column_data:
            continue # Пропустить пустой столбец
        
        #Настройка ширины столбца
        set_column_width(sheet, col, 7)
        
        max_value = max(column_data)
        
        for row in range(start_row, start_row + data_rows):
            if sheet.cell(row=row, column=col).value is None:
                continue # Пропустить пустую ячейку
            cell_value = float(sheet.cell(row=row, column=col).value)
            
            # Условное форматирование градиентом
            fill = PatternFill(start_color='FF9999', end_color='00FF00', fill_type='solid')
                        
            # Выделение максимального значения
            if cell_value == max_value:
                # Выделение максимального значения "Жирным" + размер шрифта
                sheet.cell(row=row, column=col).font = Font(bold=True, size=12)
                # Выделение максимального значения "Красным" 
                sheet.cell(row=row, column=col).fill = fill
                max_regim_cell_row = row
                
        # Вставка максимального значения под столбцом
        max_value_cell = sheet.cell(row=start_row + data_rows + 1, column=col)
        max_value_cell.value = max_value
        # Вставка режима на котором достигается максимальное значение
        max_regim_cell = sheet.cell(row=start_row + data_rows + 2, column=col)
        max_regim_cell.value = sheet.cell(row=max_regim_cell_row, column=1).value
        
        change_text_orientation(sheet, max_regim_cell.coordinate)
          
    # Сохранение изменений
    wb.save(path)
    wb.close()


#_______________________________________________________________________________________
#_______________________________________________________________________________________

# Функция для оформления ячеек границы и цвет границ
def highlight_non_empty_cells(path, sheet_number):
    wb = load_workbook(path)
    sheet = wb.worksheets[sheet_number - 1]
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None: #thick
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color='000000'),
                                                     right=openpyxl.styles.Side(style='thin', color='000000'),
                                                     top=openpyxl.styles.Side(style='thin', color='000000'),
                                                     bottom=openpyxl.styles.Side(style='thin', color='000000'))
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    wb.save(path)



def highlight_non_empty_cells_2(path, sheet_number):
    wb = load_workbook(path)
    sheet = wb.worksheets[sheet_number - 1]
    
    num_0 = 13
    
    if sheet_number == 3 or sheet_number == 6:
        num_0 = 6
    elif sheet_number == 4:
        num_0 = 4
    elif sheet_number == 5:
        num_0 = 3
    # Определяем максимальное количество строк и столбцов
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Вставляем формулу в шестой столбец для каждой строки
    for row in range(1, max_row+1):
        for col in range(2, max_col+1):
            if col % num_0 == 0:
                cell_1 = sheet.cell(row=row, column=col + 2)
                cell_1.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color='000000'),
                                                       right=openpyxl.styles.Side(style='thin', color='000000'),
                                                       top=openpyxl.styles.Side(style='thin', color='000000'),
                                                       bottom=openpyxl.styles.Side(style='thin', color='000000'))
                cell_2 = sheet.cell(row=row, column=2)
                cell_2.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thick', color='000000'),
                                                       right=cell_2.border.right,
                                                       top=cell_2.border.top,
                                                       bottom=cell_2.border.bottom)
                cell_0 = sheet.cell(row=row, column=col + 2)
                cell_0.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thick', color='000000'),
                                                       right=cell_2.border.right,
                                                       top=cell_2.border.top,
                                                       bottom=cell_2.border.bottom)
                
        for col in range(2, max_col+1):
            cell_3 = sheet.cell(row=3, column=col)
            cell_3.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thick', color='000000'),
                                                   left=cell_3.border.left,
                                                   right=cell_3.border.right,
                                                   top=cell_3.border.top)
    wb.save(path)    

# Функция объединения ячеек
def merge_cells(sheet, start_row, start_col, end_row, end_col):
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

#_______________________________________________________________________________________
#_______________________________________________________________________________________    

# Функция автоподбора ширины ячейки
def autofit_column(sheet, col):
    max_length = 0
    for row in sheet.iter_rows(values_only=True):
        for cell in row[col - 1:col]:
            try:
                if len(str(cell)) > max_length:
                    max_length = len(cell)
            except:
                pass
    sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length

#_______________________________________________________________________________________
#_______________________________________________________________________________________    
    
# Функция форматирования ячеек по заданной ширине
def set_column_width(sheet, col, width):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

#_______________________________________________________________________________________
#_______________________________________________________________________________________
   
# Функция по вертикальному выравниванию содержимого в ячейке   
def change_text_orientation(sheet, cell):
    cell_obj = sheet[cell]
    cell_obj.alignment = openpyxl.styles.Alignment(textRotation=180, horizontal="center", vertical="top")   

#_______________________________________________________________________________________
#_______________________________________________________________________________________

# Функция которая берет значения с листа отборы, выбирает максимальное Gv
# и переносит данные по соответствующему отбору на данном режиме
def copy_max_values_(
    file_path: str,
    src_sheet_name: str,
    src_start_row: int,
    src_start_col: int,
    dst_sheet_name: str,
    dst_start_row: int,
    dst_start_col: int
):
    # Открываем Excel файл
    wb = openpyxl.load_workbook(file_path)

    # Выбираем листы
    src_sheet = wb[src_sheet_name]
    dst_sheet = wb[dst_sheet_name]

    # Ищем максимальные значения в каждом 6-м столбце начиная с src_start_col
    max_values = {}
    max_rows = {}

    # Проходим по каждому 6-му столбцу начиная с src_start_col
    for col in range(src_start_col, src_sheet.max_column + 1, 6):
        max_value = float('-inf')
        max_row = None

        # Проходим по всем строкам начиная с src_start_row
        for row in range(src_start_row, src_sheet.max_row + 1):
            value = src_sheet.cell(row=row, column=col).value
            if value is not None and isinstance(value, (int, float)):
                if value > max_value:
                    max_value = value
                    max_row = row

        max_values[col] = max_value
        max_rows[col] = max_row

    # Собираем данные для копирования
    data_to_copy = []

    for col, max_row in max_rows.items():
        if max_row is not None:
            # Список для текущей строки
            row_data = []
            # Добавляем значение из первой строки (над найденным максимальным значением)
            row_data.append(src_sheet.cell(row=1, column=col-5).value)
            # Добавляем шесть значений слева от текущего максимального значения (включая само максимальное)
            for offset in range(-5, 1):
                row_data.append(src_sheet.cell(row=max_row, column=col + offset).value)
            # Добавляем список в общий список данных
            data_to_copy.append(row_data)

    # Вставляем данные на указанный лист начиная с dst_start_row и dst_start_col
    for i, row_data in enumerate(data_to_copy):
        for j, value in enumerate(row_data):
            dst_sheet.cell(row=dst_start_row + i, column=dst_start_col + j, value=value)

    # Сохраняем изменения в файле
    wb.save(file_path)
    wb.close()

#_______________________________________________________________________________________
#_______________________________________________________________________________________

import math

def select_diameter(D, G, P, V):

    """
    Функция возвращает список диаметров, удовлетворяющих заданному условию C_target для Z от 1 до 4.

    Args:
        D: Список возможных диаметров.
        G: Расход.
        P: Давление.
        V: Удельный объем пара.

    Returns:
        Список диаметров, удовлетворяющих условию C_target для Z от 1 до 4.
    """

    if 340.0 > P > 9.999999:
        C_target = 50.0
        max_diameter = 400
    elif 9.999999 > P > 2.999999:
        C_target = 70.0  # Выбрано среднее значение диапазона 60-75
        max_diameter = 800
    elif 2.999999 > P > 0.000001:
        C_target = 82.5  # Выбрано среднее значение диапазона 75-90
        max_diameter = 1400
    else:
        return (100, 1, [0, 0, 0, 0], [0, 0, 0, 0])  # Некорректное значение давления

    D_definitions = [0, 0, 0, 0]
    D_definitions_2 = [0, 0, 0, 0]

    for Z in range(1, 5):  # Итерация по Z от 1 до 4
        for diameter in D:
            if diameter > max_diameter:
                break  # Прекращаем проверку диаметров, если превышен max_diameter
            F_ = math.pi * Z * (diameter / 10) ** 2 / 4
            C = 10000 * (G / 3.6) * V / F_

            if C < C_target:  # Проверка условия C_target
                D_definitions[Z - 1] = diameter
                D_definitions_2[Z - 1] = f'{Z}/{diameter}/{round(F_, 1)}/{round(C, 1)}'
                break  # Выход из цикла по diameter, так как найдено подходящее значение для текущего Z

    if D_definitions[0] > 0:
        D = D_definitions[0]
        Z = 1
    elif D_definitions[1] > 0:
        D = D_definitions[1]
        Z = 2
    elif D_definitions[2] > 0:
        D = D_definitions[2]
        Z = 3
    elif D_definitions[3] > 0:
        D = D_definitions[3]
        Z = 4
    else:
        D = 100
        Z = 1

    return D, Z, D_definitions, D_definitions_2


def add_excel_formula(path, sheet_number, start_coord):
    # Открываем excel файл
    wb = openpyxl.load_workbook(path)
    sheet = wb.worksheets[sheet_number - 1]
    # Определяем максимальное количество строк и столбцов
    max_row = sheet.max_row

    # Вставляем формулу в шестой столбец для каждой строки
    start_row, start_col = start_coord

    # Значения диаметров СТ-255736 Dy=10-400 t<=400 Р<=22
    # Значения диаметров СТ-260140 Dy=500-1400 t<=350 P<=16
    D_list = [10, 15, 20, 25, 32, 40, 50, 65, 80, 100, 125, 150, 200, 250, 300, 350, 400, 500, 600, 700, 800, 900, 1000, 1200, 1400]

    val3 = 0  # Толщина стенки трубопровода
    val4 = 500  # Входной радиус щели
    val5 = 500  # Выходной радиус щели
    val6 = 60  # заданная скорость в щели

    for row in range(start_row, max_row + 1):
        cell4 = sheet.cell(row=row, column=2)  # расход
        cell11 = sheet.cell(row=row, column=4)  # давление
        cell5 = sheet.cell(row=row, column=6)  # удельный объем

        select_diameter_res = select_diameter(D_list, cell4.value, cell11.value, cell5.value)
        val2 = select_diameter_res[0]
        val1 = select_diameter_res[1]
        diameter_list = select_diameter_res[3]


        sheet.cell(row=row, column=21, value=f'{diameter_list[0]}')
        sheet.cell(row=row, column=22, value=f'{diameter_list[1]}')
        sheet.cell(row=row, column=23, value=f'{diameter_list[2]}')
        sheet.cell(row=row, column=24, value=f'{diameter_list[3]}')

        sheet.cell(row=row, column=9, value=f'={val1}')
        sheet.cell(row=row, column=10, value=f'={val2}')
        sheet.cell(row=row, column=11, value=f'={val3}')
        sheet.cell(row=row, column=15, value=f'={val4}')
        sheet.cell(row=row, column=16, value=f'={val5}')
        sheet.cell(row=row, column=17, value=f'={val6}')

        cell1 = sheet.cell(row=row, column=9)
        cell2 = sheet.cell(row=row, column=10)
        cell3 = sheet.cell(row=row, column=11)
        cell6 = sheet.cell(row=row, column=12)

        # Расчет площади трубопровода отбора
        formula1 = f'=3.1415924565*{cell1.coordinate}*(({cell2.coordinate}/10)-2*({cell3.coordinate}/10))^2/4'
        sheet.cell(row=row, column=12, value=formula1)
        # Расчет скорости в трубопроводе отбора
        formula2 = f'=10000*({cell4.coordinate}/3.6)*{cell5.coordinate}/{cell6.coordinate}'
        sheet.cell(row=row, column=13, value=formula2)
        cell7 = sheet.cell(row=row, column=15) # R1
        cell8 = sheet.cell(row=row, column=16) # R2
        cell9 = sheet.cell(row=row, column=17) # Cщели
        formula3 = f'=10000*({cell4.coordinate}/3.6)*{cell5.coordinate}/{cell9.coordinate}'
        sheet.cell(row=row, column=18, value=formula3)
        cell10 = sheet.cell(row=row, column=18)
        formula4 = f'=({cell10.coordinate}*100)/(3.1415924565*({cell7.coordinate}+{cell8.coordinate}))'
        sheet.cell(row=row, column=19, value=formula4)


    # Сохраняем изменения в файле
    wb.save(path)
    wb.close()

#_______________________________________________________________________________________
#_______________________________________________________________________________________


# Функция для обработки выбора папки
def choose_folder():
    folder_path = filedialog.askdirectory() # Открываем диалог выбора папки
    save_button.config(state=tk.NORMAL) # Делаем кнопку сохранить активной
    global selected_folder
    selected_folder = folder_path

# Функция для обработки кнопки сохранить
def save_handler():
    process_files(selected_folder)

# Функция для обработки файлов в выбранной папке
def process_files(folder_path):
    save_path = filedialog.askdirectory() # открываем диалог выбора папки для сохранения
    new_workbook = openpyxl.Workbook() # Создаем новый excel файл
    main_sheet = new_workbook.active
    main_sheet.title = "Сводная"
    compartment_sheet = new_workbook.create_sheet(title="Отсеки")
    extraction_sheet = new_workbook.create_sheet(title="Отборы")
    regeneration_sheet = new_workbook.create_sheet(title="Регенерация")
    water_heating_sheet = new_workbook.create_sheet(title="Сетевая вода")
    #adjustments_sheet = new_workbook.create_sheet(title="Корректировки")
    speed_in_tube_sheet = new_workbook.create_sheet(title="Скорости по отборам")
    #max_parameters_table_sheet = new_workbook.create_sheet(title="Таблица максимальных параметров")
    #max_parameters_BPR_sheet = new_workbook.create_sheet(title="ИД для расчета БПР")
    all_balances_sheet = new_workbook.create_sheet(title="Исходные данные")
    row_num = 3
    files_list = os.listdir(folder_path)
    
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')] # Перечень Excel файлов в папке
    total_files = len(excel_files)                                            # Сколько всего Excel файлов в папке
    copied_files = 0                                                          # Переменная для вычисления сколько файлов копировано
    row_all_balances = 1
    for file_name in os.listdir(folder_path):                                 # Перебираем файлы в выбранной папке
        if file_name.endswith('.xlsx'):                                       # Проверяем, что файл - это Excel
                file = folder_path + '/' + file_name                          # Переменная в которой хранится путь к конкретному Excel файлу
                
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                
                # Определяем тип баланса и его переменные
                balance_type, name_extraction, name_regeneration, name_water_heating, compartment_name_value_list, compartment_name_unit_list, main_parameters_list, main_parameters_value_list, main_parameters_unit_list, type_balance_regeneration = type_balance(file)
                
                # Определение областей
                start_coord = (11, 3)
                empty_rows = find_empty_rows(file, start_coord)
                
                # Область отсеков
                cell_compartment_start = find_words_in_excel(file, name_compartment, 2, 0)
                cell_compartment_end = find_words_in_excel(file, name_extraction, -1, 13)
                cell_compartment_range = cell_compartment_start.coordinate + ':' + cell_compartment_end.coordinate

                # Область отборов
                cell_extraction_start = find_words_in_excel(file, name_extraction, 1, 0)
                cell_extraction_end = find_words_in_excel(file, name_regeneration, -1, 13)
                cell_extraction_range = cell_extraction_start.coordinate + ':' + cell_extraction_end.coordinate

                # Область регенерации
                cell_regeneration_start = find_words_in_excel(file, name_regeneration, 1, 0)
                cell_regeneration_row_end = empty_rows[2] - 1
                cell_regeneration_end = find_words_in_excel_2(file, name_compartment, cell_regeneration_row_end, 11)
                cell_regeneration_range = cell_regeneration_start.coordinate + ':' + cell_regeneration_end.coordinate
                                
                # Область подогрева сетевой воды
                cell_water_heating_start = new_cell_to_excel(file, cell_regeneration_end.coordinate, 2, -11)
                cell_water_heating_start_2 = new_cell_to_excel(file, cell_regeneration_end.coordinate, 2, -10)
                cell_water_heating_start_value = sheet[cell_water_heating_start].value
                cell_water_heating_start_2_value = sheet[cell_water_heating_start_2].value
                if cell_water_heating_start_value is not None and cell_water_heating_start_2_value is not None:
                    cell_water_heating_start = new_cell_to_excel(file, cell_regeneration_end.coordinate, 2, -11)
                    cell_water_heating_end = new_cell_to_excel(file, cell_water_heating_start, 0, 11)
                    cell_water_heating_range = cell_water_heating_start + ':' + cell_water_heating_end
                else:
                    cell_water_heating_start = new_cell_to_excel(file, cell_regeneration_end.coordinate, 2, -10)
                    cell_water_heating_end = new_cell_to_excel(file, cell_water_heating_start, 0, 11)
                    cell_water_heating_range = cell_water_heating_start + ':' + cell_water_heating_end
                
                header_row = 1
                
#ТАБЛИЦА ОСНОВНЫХ ПАРАМЕТРОВ
                # Запускаем функцию сбора основных параметров
                main_parameters_list_coordinate = []
                for i in main_parameters_list_all:
                    cell_value = find_words_in_excel(file, i, 1, 0).coordinate
                    main_parameters_list_coordinate.append(cell_value)
                main_parameters_list = main_parameters_list_coordinate + main_parameters_list
                result_0 = process_excel_main(file, main_parameters_list)     
                
                # Записываем данные в таблицу Excel
                for i, parameter in enumerate(result_0):
                    main_sheet.cell(row=row_num+1, column=i+2, value=parameter)
                
                # Записываем название файла в первый столбец
                main_sheet.cell(row=row_num+1, column=1, value=file_name)
                
                # Выводим наименования и единицы измерения параметров
                main_parameters_value_list = main_parameters_value_list_all + main_parameters_value_list
                for i, value_unit in enumerate(main_parameters_value_list):
                    main_sheet.cell(row=2, column=i+2, value=value_unit)
                main_parameters_unit_list = main_parameters_unit_list_all + main_parameters_unit_list
                for i, value_unit in enumerate(main_parameters_unit_list):
                    main_sheet.cell(row=3, column=i+2, value=value_unit)    
          
                
#ТАБЛИЦА ОТСЕКОВ                
                # Запускаем функцию сбора параметров отсеков
                result_1 = process_excel_compartment(file, cell_compartment_start.row, cell_compartment_end.row, cell_compartment_start.column)
                # Записываем название файла в первый столбец
                compartment_sheet.cell(row=row_num+1, column=1, value=file_name) # Записываем название файла
                # Выводим ключи в первую строку
                for col, key in enumerate(result_1.keys()):
                    compartment_sheet.cell(row=header_row, column=col*len(compartment_name_value_list)+1+1, value=key)
                    # Объединяем ячейки отсеков
                    merge_cells(compartment_sheet, header_row, col*len(compartment_name_value_list)+1+1, header_row, (col+1)*len(compartment_name_value_list)+1)
                    
                # Выводим наименования и единицы измерения параметров
                for i, (key, values) in enumerate(result_1.items(), start=0):
                    for j in range(len(compartment_name_value_list)):
                        #Вывод наименование параметров
                        compartment_sheet.cell(row=header_row+1, column=j+2+i*len(compartment_name_value_list), value=compartment_name_value_list[j])
                        #Вывод наименование размерности
                        compartment_sheet.cell(row=header_row+2, column=j+2+i*len(compartment_name_value_list), value=compartment_name_unit_list[j])
                
                result_row = 1
                # Выводим значения по каждому ключу в строку под ним
                for i, (key, values) in enumerate(result_1.items(), start=0):
                    for col, value in enumerate(values):
                        compartment_sheet.cell(row=result_row+row_num, column=col+1+i*len(compartment_name_value_list)+1, value=value)
                row_num += 1

#ТАБЛИЦА ОТБОРОВ
                # Запускаем функцию сбора параметров отборов
                result_2 = process_excel_extraction(file, cell_extraction_range)
                col_num = 2
                # Записываем название файла в первый столбец
                extraction_sheet.cell(row=row_num, column=1, value=file_name)
                # Записываем данные в таблицу Excel
                for key in result_2:
                    extraction_sheet.cell(row=header_row, column=col_num, value=key)
                    # Объединяем ячейки отборов
                    merge_cells(extraction_sheet, header_row, col_num, header_row, col_num+5)
                    for sublist in result_2[key]:
                        extraction_sheet.cell(row=header_row+1, column=col_num, value=sublist[0])
                        extraction_sheet.cell(row=row_num, column=col_num, value=sublist[1])
                        col_num += 1
                #Вывод наименование размерности
                for i, (key, values) in enumerate(result_2.items(), start=0):   
                    for j in range(len(extraction_name_unit_list)):
                        extraction_sheet.cell(row=header_row+2, column=j+2+i*len(extraction_name_unit_list), value=extraction_name_unit_list[j])
                
#ТАБЛИЦА РЕГЕНЕРАЦИИ                 
                col_num_ = 2
                # Запускаем функцию сбора параметров линии регенерации
                result_3 = process_excel_regeneration(file, cell_regeneration_range, type_balance_regeneration)
                # Записываем название файла в первый столбец
                regeneration_sheet.cell(row=row_num, column=1, value=file_name)
                # Записываем данные в таблицу Excel
                for i, (key, values) in enumerate(result_3.items(), start=0):
                    for col, value in enumerate(values):
                        regeneration_sheet.cell(row=result_row+row_num-1, column=col+1+i*4+1, value=value)
                for key in result_3:
                    regeneration_sheet.cell(row=header_row, column=col_num_, value=key)
                    for j in result_3[key]:
                        col_num_ += 1
                # Объединение ячеек регенерации
                for i, (key, values) in enumerate(result_3.items(), start=0):   
                    merge_cells(regeneration_sheet, header_row, i*4+2, header_row, (i+1)*4+1)
                #Вывод наименование размерности
                for i, (key, values) in enumerate(result_3.items(), start=0):   
                    for j in range(len(regeneration_name_value_list)):
                        regeneration_sheet.cell(row=header_row+1, column=j+2+i*len(regeneration_name_value_list), value=regeneration_name_value_list[j])    
                        regeneration_sheet.cell(row=header_row+2, column=j+2+i*len(regeneration_name_unit_list), value=regeneration_name_unit_list[j])

#ТАБЛИЦА ПОДОГРЕВА СЕТЕВОЙ ВОДЫ
                col_num_2 = 2
                # Запускаем функцию сбора параметров линии сетевой воды
                result_4 = process_excel_water_heating(file, cell_water_heating_range, 3)
                # Записываем название файла в первый столбец
                water_heating_sheet.cell(row=row_num, column=1, value=file_name)
                # Записываем данные в таблицу Excel
                for i, (key, values) in enumerate(result_4.items(), start=0):
                    for col, value in enumerate(values):
                        water_heating_sheet.cell(row=result_row+row_num-1, column=col+1+i*3+1, value=value)
                for key in result_4:
                    water_heating_sheet.cell(row=header_row, column=col_num_2, value=key)
                    for j in result_4[key]:
                        col_num_2 += 1
                # Объединение ячеек сетевой воды
                for i, (key, values) in enumerate(result_4.items(), start=0):   
                    merge_cells(water_heating_sheet, header_row, i*3+2, header_row, (i+1)*3+1)
                #Вывод наименование размерности
                for i, (key, values) in enumerate(result_4.items(), start=0):   
                    for j in range(len(water_heating_name_value_list)):
                        water_heating_sheet.cell(row=header_row+1, column=j+2+i*len(water_heating_name_value_list), value=water_heating_name_value_list[j])    
                        water_heating_sheet.cell(row=header_row+2, column=j+2+i*len(water_heating_name_unit_list), value=water_heating_name_unit_list[j])

# ТАБЛИЦА СКОРОСТИ ПО ОТБОРАМ

                # Вывод наименование размерности
                for i, (key, values) in enumerate(result_2.items(), start=0):
                    for j in range(len(extraction_name_unit_list)):
                        speed_in_tube_sheet.cell(row=header_row + 2, column=j + 2, value=extraction_name_unit_list[j])
                        speed_in_tube_sheet.cell(row=header_row + 1, column=j + 2, value=extraction_name_value_list[j])
                    for a in range(len(speed_name_value_list)):
                        speed_in_tube_sheet.cell(row=header_row + 2, column=a + 8, value=speed_name_unit_list[a])
                        speed_in_tube_sheet.cell(row=header_row + 1, column=a + 8, value=speed_name_value_list[a])


#ТАБЛИЦА ИСХОДНЫХ ДАННЫХ
                # Сохранение баланса целиком на лист балансы
                for row_ in sheet.iter_rows(values_only=True):
                    all_balances_sheet.append(row_)
                    max_row = sheet.max_row
                    
                # Гиперссылки на балансы из каждого листа
                hyperlink = f"#'{all_balances_sheet.title}'!A{row_all_balances}"
                main_sheet[f"A{row_num}"].hyperlink = hyperlink
                compartment_sheet[f"A{row_num}"].hyperlink = hyperlink
                extraction_sheet[f"A{row_num}"].hyperlink = hyperlink
                regeneration_sheet[f"A{row_num}"].hyperlink = hyperlink
                water_heating_sheet[f"A{row_num}"].hyperlink = hyperlink
                
                row_all_balances += max_row
                
                save_path_1 = save_path + '/Сводка.xlsx'      # Формируем путь для нового файла excel
                new_workbook.save(save_path_1)                # Сохраняем в новый файл excel
                new_workbook.close()                          # закрываем новый excel файл
                
                # вывод в окно интерфейса сколько файлов скопировано
                copied_files +=1
                status_label.config(text=f'Копировано файлов {copied_files} из {total_files}')




                
                # обновление окна интерфейса каждый цикл
                root.update()
        

    # Пост-обработка сводки
    status_label.config(text=f'Раскрашиваем, анализируем, смотрим...')
    root.update()
    highlight_non_empty_cells(save_path_1, 1)
    process_excel_file_1(save_path_1, 1, (4, 2))

    # Добавление данных в таблицу расчета скоростей
    copy_max_values_(save_path_1, "Отборы",4, 7, "Скорости по отборам", 4, 1)
    add_excel_formula(save_path_1, 6, (4, 2))

    highlight_non_empty_cells(save_path_1, 2)
    highlight_non_empty_cells_2(save_path_1, 2)
    process_excel_file(save_path_1, 2, (4, 2))
    
    highlight_non_empty_cells(save_path_1, 3)
    highlight_non_empty_cells_2(save_path_1, 3)
    process_excel_file(save_path_1, 3, (4, 2))
    
    highlight_non_empty_cells(save_path_1, 4)
    highlight_non_empty_cells_2(save_path_1, 4)
    process_excel_file(save_path_1, 4, (4, 2))
    
    highlight_non_empty_cells(save_path_1, 5)
    highlight_non_empty_cells_2(save_path_1, 5)
    process_excel_file(save_path_1, 5, (4, 2))

    highlight_non_empty_cells(save_path_1, 6)



    status_label.config(text=f'Готово!')
    
                    
# Создаем GUI окно
root = tk.Tk()
root.title("VP_v1.06")
root.geometry("300x200")

# Инструкция к кнопке "Выбор папки"
status_label_0 = tk.Label(root, text='Выберете папку с балансами и зайдите в нее')
status_label_0.pack()

# Кнопка для выбора папки
choose_button = tk.Button(root, text="Выбрать папку...", command=choose_folder)
choose_button.pack(pady=10)

# Инструкция к кнопке "Сохранить в..."
status_label_0 = tk.Label(root, text='Выберете папку куда будет сохранена сводка')
status_label_0.pack()

# Кнопка для сохранения
save_button = tk.Button(root, text="Сохранить в...", command=save_handler, state=tk.DISABLED)
save_button.pack(pady=10)

status_label = tk.Label(root, text='')
status_label.pack()

root.mainloop()